"""Exportación de comprobantes electrónicos a Excel, para el contador.

El formato sigue la lógica del **registro de ventas**: una fila por documento
con el desglose de base imponible e IGV, los códigos del catálogo SUNAT y el
estado real ante SUNAT, más una hoja de resumen que cuadra los totales del
periodo.
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.comprobante import (
    ETIQUETAS_ESTADO_COMPROBANTE,
    ETIQUETAS_TIPO_COMPROBANTE,
    ComprobanteElectronico,
    EstadoComprobante,
)
from app.services.factpro_catalogos import TIPO_COMPROBANTE_SUNAT
from app.services.facturacion import desglosar_igv
from app.services.ficha_pdf import EMPRESA

ROJO = "E2201F"
NEGRO = "111111"
GRIS = "F2F2F2"

#: Estados que sí suman en el registro de ventas. Un rechazado o un error nunca
#: llegó a existir para SUNAT, y un anulado se compensa con su nota de crédito.
ESTADOS_VALIDOS = {EstadoComprobante.ACEPTADO, EstadoComprobante.REGISTRADO}

COLUMNAS = [
    ("Fecha emisión", 14),
    ("Cód. SUNAT", 11),
    ("Tipo", 16),
    ("Serie", 9),
    ("Número", 10),
    ("Tipo doc.", 10),
    ("N° documento", 15),
    ("Cliente / razón social", 38),
    ("Moneda", 8),
    ("Base imponible", 15),
    ("IGV (18%)", 13),
    ("Total", 13),
    ("Estado", 13),
    ("Estado SUNAT", 22),
    ("Venta", 12),
    ("Emitido por", 22),
    ("Hash CPE", 26),
    ("Observación", 30),
]


def _montos(c: ComprobanteElectronico) -> tuple[Decimal, Decimal, Decimal] | None:
    """Devuelve (base, igv, total), o None si el importe no es recuperable.

    Se leen del propio comprobante, congelados al emitirlo. La venta sólo se
    usa como respaldo para los documentos anteriores a ese cambio; si tampoco
    está, la fila sale con las celdas de dinero vacías. Escribir 0.00 sería
    peor que dejarlo en blanco: en un registro de ventas ese cero se suma y
    subdeclara el periodo.
    """
    if c.total is not None:
        base = c.base_imponible if c.base_imponible is not None else Decimal("0.00")
        igv = c.igv if c.igv is not None else Decimal("0.00")
        return base, igv, c.total
    if c.venta is not None:
        return desglosar_igv(Decimal(c.venta.total))
    return None


def _observacion(c: ComprobanteElectronico, sin_importe: bool) -> str:
    partes = []
    if sin_importe:
        partes.append("Importe no recuperable: la venta asociada ya no existe")
    if c.es_simulado:
        partes.append("SIMULADO — sin validez ante SUNAT")
    if c.motivo_anulacion:
        partes.append(f"Anulado: {c.motivo_anulacion}")
    if c.mensaje_error:
        partes.append(c.mensaje_error[:120])
    return " · ".join(partes)


def _cabecera(ws, fila: int) -> None:
    fill = PatternFill("solid", fgColor=NEGRO)
    fuente = Font(bold=True, color="FFFFFF", size=9)
    for i, (nombre, ancho) in enumerate(COLUMNAS, start=1):
        celda = ws.cell(row=fila, column=i, value=nombre)
        celda.fill = fill
        celda.font = fuente
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = ancho
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)


def exportar_comprobantes_excel(
    comprobantes: list[ComprobanteElectronico],
    desde: date,
    hasta: date,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro de ventas"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS))
    titulo = ws.cell(row=1, column=1, value=f"{EMPRESA['razon_social'].upper()} — REGISTRO DE VENTAS")
    titulo.font = Font(bold=True, size=14, color=ROJO)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNAS))
    ws.cell(
        row=2,
        column=1,
        value=(
            f"RUC {EMPRESA['ruc']} · Periodo del {desde.strftime('%d/%m/%Y')} "
            f"al {hasta.strftime('%d/%m/%Y')} · {len(comprobantes)} documento(s)"
        ),
    ).font = Font(italic=True, size=10)

    fila = 4
    _cabecera(ws, fila)
    fila += 1

    resumen: dict[str, dict[str, Decimal | int]] = {}
    hay_simulados = False

    sin_importe = 0

    for c in comprobantes:
        montos = _montos(c)
        base, igv, total = montos if montos else (None, None, None)
        anulado = c.estado is EstadoComprobante.ANULADO
        hay_simulados = hay_simulados or c.es_simulado
        if montos is None:
            sin_importe += 1

        valores = [
            c.fecha_emision,
            TIPO_COMPROBANTE_SUNAT.get(c.tipo.value, ""),
            ETIQUETAS_TIPO_COMPROBANTE[c.tipo.value],
            c.serie,
            c.numero,
            c.cliente_tipo_documento,
            c.cliente_numero_documento,
            c.cliente_denominacion,
            c.moneda,
            float(base) if base is not None else None,
            float(igv) if igv is not None else None,
            float(total) if total is not None else None,
            ETIQUETAS_ESTADO_COMPROBANTE[c.estado.value],
            c.descripcion_estado_sunat or "",
            c.venta.numero if c.venta else "",
            c.usuario.full_name if c.usuario else "",
            c.hash_cpe or "",
            _observacion(c, montos is None),
        ]
        for i, v in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=i, value=v)
            celda.font = Font(size=9)
            if i == 1:
                celda.number_format = "DD/MM/YYYY"
            if i in (10, 11, 12):
                celda.number_format = "#,##0.00"
            # Lo anulado se tacha en vez de esconderse: el contador necesita
            # ver el correlativo completo, sin huecos que parezcan documentos
            # perdidos.
            if anulado:
                celda.font = Font(size=9, strike=True, color="888888")

        if montos is not None and c.estado in ESTADOS_VALIDOS:
            acc = resumen.setdefault(
                c.tipo.value,
                {"cantidad": 0, "base": Decimal("0.00"), "igv": Decimal("0.00"), "total": Decimal("0.00")},
            )
            acc["cantidad"] = int(acc["cantidad"]) + 1
            acc["base"] = Decimal(acc["base"]) + base
            acc["igv"] = Decimal(acc["igv"]) + igv
            acc["total"] = Decimal(acc["total"]) + total

        fila += 1

    # ------------------------------------------------------------------ Resumen
    hoja = wb.create_sheet("Resumen")
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    hoja.cell(row=1, column=1, value="RESUMEN DEL PERIODO").font = Font(bold=True, size=13, color=ROJO)
    hoja.cell(
        row=2,
        column=1,
        value=f"Del {desde.strftime('%d/%m/%Y')} al {hasta.strftime('%d/%m/%Y')}",
    ).font = Font(italic=True, size=10)

    f = 4
    for i, h in enumerate(["Tipo de comprobante", "Cantidad", "Base imponible", "IGV (18%)", "Total"], start=1):
        celda = hoja.cell(row=f, column=i, value=h)
        celda.fill = PatternFill("solid", fgColor=NEGRO)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center")
        hoja.column_dimensions[get_column_letter(i)].width = 24 if i == 1 else 16
    f += 1

    totales = {"cantidad": 0, "base": Decimal("0.00"), "igv": Decimal("0.00"), "total": Decimal("0.00")}
    for tipo, acc in sorted(resumen.items()):
        hoja.cell(row=f, column=1, value=ETIQUETAS_TIPO_COMPROBANTE[tipo])
        hoja.cell(row=f, column=2, value=int(acc["cantidad"]))
        for col, clave in ((3, "base"), (4, "igv"), (5, "total")):
            celda = hoja.cell(row=f, column=col, value=float(Decimal(acc[clave])))
            celda.number_format = "#,##0.00"
        totales["cantidad"] = int(totales["cantidad"]) + int(acc["cantidad"])
        for clave in ("base", "igv", "total"):
            totales[clave] = Decimal(totales[clave]) + Decimal(acc[clave])
        f += 1

    hoja.cell(row=f, column=1, value="TOTAL").font = Font(bold=True)
    hoja.cell(row=f, column=2, value=int(totales["cantidad"])).font = Font(bold=True)
    for col, clave in ((3, "base"), (4, "igv"), (5, "total")):
        celda = hoja.cell(row=f, column=col, value=float(Decimal(totales[clave])))
        celda.number_format = "#,##0.00"
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor=GRIS)

    f += 2
    hoja.cell(
        row=f,
        column=1,
        value="Sólo suman los comprobantes aceptados o registrados; los anulados, rechazados y con error se listan tachados pero no se totalizan.",
    ).font = Font(italic=True, size=9)

    if sin_importe:
        f += 1
        alerta = hoja.cell(
            row=f,
            column=1,
            value=(
                f"ATENCIÓN: {sin_importe} documento(s) quedaron sin importe porque su venta "
                "asociada ya no existe en el sistema; revísalos antes de declarar."
            ),
        )
        alerta.font = Font(bold=True, size=9, color=ROJO)

    if hay_simulados:
        f += 1
        aviso = hoja.cell(
            row=f,
            column=1,
            value="ATENCIÓN: el periodo incluye comprobantes SIMULADOS (emitidos sin conexión real a SUNAT). No son documentos tributarios válidos.",
        )
        aviso.font = Font(bold=True, size=9, color=ROJO)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
