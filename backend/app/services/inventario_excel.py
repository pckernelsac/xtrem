"""Importación de productos desde Excel.

Diseñada para que el error sea barato: por defecto corre en modo prueba y
devuelve el reporte fila por fila sin escribir nada. Recién con
`modo_prueba=False` aplica los cambios, y ahí lo hace todo o nada.
"""

import uuid
from decimal import Decimal, InvalidOperation
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.inventario import Categoria, Producto, TipoMovimiento, UnidadMedida
from app.schemas.inventario import FilaImportacion, ResultadoImportacion
from app.services.inventario import registrar_movimiento

#: Encabezado esperado -> atributo del modelo. Se aceptan variantes comunes
#: para no obligar a nadie a escribir la cabecera con tildes exactas.
COLUMNAS: dict[str, str] = {
    "sku": "sku",
    "codigo": "sku",
    "código": "sku",
    "nombre": "nombre",
    "descripcion": "descripcion",
    "descripción": "descripcion",
    "marca": "marca",
    "categoria": "categoria",
    "categoría": "categoria",
    "unidad": "unidad",
    "stock": "stock",
    "stock actual": "stock",
    "stock minimo": "stock_minimo",
    "stock mínimo": "stock_minimo",
    "minimo": "stock_minimo",
    "mínimo": "stock_minimo",
    "precio compra": "precio_compra",
    "costo": "precio_compra",
    "precio venta": "precio_venta",
    "precio": "precio_venta",
    "codigo de barras": "codigo_barras",
    "código de barras": "codigo_barras",
    "ubicacion": "ubicacion",
    "ubicación": "ubicacion",
}

OBLIGATORIAS = {"sku", "nombre"}


def _texto(valor) -> str | None:
    if valor is None:
        return None
    s = str(valor).strip()
    return s or None


def _decimal(valor, campo: str) -> Decimal:
    if valor is None or str(valor).strip() == "":
        return Decimal("0")
    try:
        # Se acepta la coma decimal: es lo que sale de un Excel en español.
        return Decimal(str(valor).strip().replace(",", ".").replace("S/", "").strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"'{campo}' no es un número válido: {valor!r}") from exc


def plantilla_xlsx() -> bytes:
    """Excel de ejemplo con las cabeceras que espera el importador."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    cabeceras = [
        "SKU", "Nombre", "Descripcion", "Marca", "Categoria", "Unidad",
        "Stock", "Stock minimo", "Precio compra", "Precio venta",
        "Codigo de barras", "Ubicacion",
    ]
    ws.append(cabeceras)
    ws.append([
        "CAD-XT-12V", "Cadena 12v XT M8100", "Cadena de 126 eslabones", "Shimano",
        "Transmisión", "UNIDAD", 12, 3, 145.00, 189.90, "7891234567890", "Estante A-3",
    ])
    ws.append([
        "PAS-RES-B03", "Pastillas de freno B03S resina", "", "Shimano",
        "Frenos", "PAR", 25, 8, 28.00, 45.00, "", "Estante B-1",
    ])

    for i, ancho in enumerate([16, 34, 30, 14, 16, 10, 8, 12, 14, 13, 18, 14], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _leer_cabeceras(ws) -> dict[str, int]:
    """Mapea atributo -> índice de columna, tolerando mayúsculas y tildes."""
    mapa: dict[str, int] = {}
    for idx, celda in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
        clave = (str(celda).strip().lower() if celda is not None else "")
        if clave in COLUMNAS:
            mapa[COLUMNAS[clave]] = idx

    faltan = OBLIGATORIAS - mapa.keys()
    if faltan:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"Al archivo le faltan columnas obligatorias: {', '.join(sorted(faltan))}. "
                "Descarga la plantilla desde el botón 'Plantilla Excel'."
            ),
        )
    return mapa


def importar_productos(
    db: Session,
    contenido: bytes,
    usuario_id: uuid.UUID | None,
    modo_prueba: bool = True,
) -> ResultadoImportacion:
    try:
        wb = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="El archivo no es un Excel válido (.xlsx)",
        ) from exc

    ws = wb.active
    mapa = _leer_cabeceras(ws)

    filas: list[FilaImportacion] = []
    creados = actualizados = errores = 0

    # Caché de categorías por nombre en minúsculas: evita una consulta por fila
    # y permite que varias filas creen y reutilicen la misma categoría nueva.
    cache_categorias: dict[str, Categoria] = {
        c.nombre.lower(): c for c in db.scalars(select(Categoria)).all()
    }
    # SKUs ya vistos en este archivo, para detectar duplicados internos.
    vistos: set[str] = set()

    for n, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if fila is None or all(c is None or str(c).strip() == "" for c in fila):
            continue

        def val(campo: str):
            idx = mapa.get(campo)
            return fila[idx] if idx is not None and idx < len(fila) else None

        sku = _texto(val("sku"))
        try:
            if not sku:
                raise ValueError("El SKU está vacío")
            sku = sku.upper().replace(" ", "")

            if sku in vistos:
                raise ValueError("SKU repetido dentro del mismo archivo")
            vistos.add(sku)

            nombre = _texto(val("nombre"))
            if not nombre:
                raise ValueError("El nombre está vacío")

            unidad_txt = (_texto(val("unidad")) or "UNIDAD").upper()
            try:
                unidad = UnidadMedida(unidad_txt)
            except ValueError as exc:
                validas = ", ".join(u.value for u in UnidadMedida)
                raise ValueError(f"Unidad '{unidad_txt}' inválida. Usa: {validas}") from exc

            stock = _decimal(val("stock"), "Stock")
            if stock < 0:
                raise ValueError("El stock no puede ser negativo")

            categoria = None
            nombre_cat = _texto(val("categoria"))
            if nombre_cat:
                categoria = cache_categorias.get(nombre_cat.lower())
                if categoria is None:
                    categoria = Categoria(nombre=nombre_cat)
                    db.add(categoria)
                    db.flush()
                    cache_categorias[nombre_cat.lower()] = categoria

            datos = {
                "nombre": nombre,
                "descripcion": _texto(val("descripcion")),
                "marca": _texto(val("marca")),
                "categoria_id": categoria.id if categoria else None,
                "unidad": unidad,
                "stock_minimo": _decimal(val("stock_minimo"), "Stock mínimo"),
                "precio_compra": _decimal(val("precio_compra"), "Precio compra"),
                "precio_venta": _decimal(val("precio_venta"), "Precio venta"),
                "codigo_barras": _texto(val("codigo_barras")),
                "ubicacion": _texto(val("ubicacion")),
            }

            existente = db.scalar(select(Producto).where(Producto.sku == sku))

            if existente:
                for campo, valor in datos.items():
                    setattr(existente, campo, valor)
                # El stock del archivo se trata como conteo físico: entra como
                # AJUSTE con su asiento, nunca pisando el saldo en silencio.
                if stock != existente.stock_actual:
                    registrar_movimiento(
                        db, existente.id, TipoMovimiento.AJUSTE, stock,
                        usuario_id=usuario_id,
                        motivo=f"Importación Excel (fila {n})",
                    )
                actualizados += 1
                filas.append(FilaImportacion(fila=n, sku=sku, accion="actualizado"))
            else:
                producto = Producto(sku=sku, stock_actual=Decimal("0"), **datos)
                db.add(producto)
                db.flush()
                if stock > 0:
                    registrar_movimiento(
                        db, producto.id, TipoMovimiento.ENTRADA, stock,
                        usuario_id=usuario_id,
                        costo_unitario=datos["precio_compra"] or None,
                        motivo=f"Carga inicial por importación (fila {n})",
                    )
                creados += 1
                filas.append(FilaImportacion(fila=n, sku=sku, accion="creado"))

        except HTTPException as exc:
            errores += 1
            filas.append(
                FilaImportacion(fila=n, sku=sku, accion="error", detalle=str(exc.detail))
            )
        except Exception as exc:
            errores += 1
            filas.append(FilaImportacion(fila=n, sku=sku, accion="error", detalle=str(exc)))

    wb.close()

    # Todo o nada: un archivo con errores no debe dejar medio inventario
    # cargado y la otra mitad sin cargar.
    if modo_prueba or errores:
        db.rollback()
    else:
        db.commit()

    return ResultadoImportacion(
        modo_prueba=modo_prueba or errores > 0,
        total_filas=len(filas),
        creados=creados,
        actualizados=actualizados,
        errores=errores,
        filas=filas,
    )
