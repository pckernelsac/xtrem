"""Exportación de reportes a Excel (openpyxl) y PDF (WeasyPrint)."""

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.fechas import hoy_local
from app.services.ficha_pdf import EMPRESA, _asset_data_url, _env

ROJO = "E2201F"
NEGRO = "111111"


def _estilar_cabecera(ws, fila: int, columnas: int) -> None:
    fill = PatternFill("solid", fgColor=NEGRO)
    fuente = Font(bold=True, color="FFFFFF")
    for col in range(1, columnas + 1):
        celda = ws.cell(row=fila, column=col)
        celda.fill = fill
        celda.font = fuente
        celda.alignment = Alignment(horizontal="center")


def _titulo(ws, texto: str, subtitulo: str, columnas: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columnas)
    c = ws.cell(row=1, column=1, value=f"ZONA XTREMA — {texto}")
    c.font = Font(bold=True, size=14, color=ROJO)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columnas)
    ws.cell(row=2, column=2)
    ws.cell(row=2, column=1, value=subtitulo).font = Font(italic=True, size=10)
    return 4  # primera fila de contenido


def exportar_ventas_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    fila = _titulo(ws, "Reporte de ventas", f"Del {data['desde']} al {data['hasta']}", 3)

    ws.cell(row=fila, column=1, value="Total vendido").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=float(data["total"]))
    ws.cell(row=fila + 1, column=1, value="Cantidad de ventas").font = Font(bold=True)
    ws.cell(row=fila + 1, column=2, value=data["cantidad"])
    ws.cell(row=fila + 2, column=1, value="Ticket promedio").font = Font(bold=True)
    ws.cell(row=fila + 2, column=2, value=float(data["ticket_promedio"]))

    fila += 4
    ws.cell(row=fila, column=1, value="VENTAS POR DÍA").font = Font(bold=True, color=ROJO)
    fila += 1
    for i, h in enumerate(["Fecha", "Ventas", "Total S/"], start=1):
        ws.cell(row=fila, column=i, value=h)
    _estilar_cabecera(ws, fila, 3)
    fila += 1
    for d in data["por_dia"]:
        ws.cell(row=fila, column=1, value=d["fecha"])
        ws.cell(row=fila, column=2, value=d["cantidad"])
        ws.cell(row=fila, column=3, value=float(d["total"]))
        fila += 1

    fila += 1
    ws.cell(row=fila, column=1, value="POR MÉTODO DE PAGO").font = Font(bold=True, color=ROJO)
    fila += 1
    for metodo, monto in data["por_metodo"].items():
        ws.cell(row=fila, column=1, value=metodo)
        ws.cell(row=fila, column=2, value=float(monto))
        fila += 1

    for col, ancho in [(1, 22), (2, 14), (3, 14)]:
        ws.column_dimensions[get_column_letter(col)].width = ancho

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_productos_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos vendidos"

    fila = _titulo(
        ws, "Productos más vendidos", f"Del {data['desde']} al {data['hasta']}", 4
    )
    for i, h in enumerate(["SKU", "Producto", "Cantidad", "Importe S/"], start=1):
        ws.cell(row=fila, column=i, value=h)
    _estilar_cabecera(ws, fila, 4)
    fila += 1
    for it in data["items"]:
        ws.cell(row=fila, column=1, value=it["sku"] or "—")
        ws.cell(row=fila, column=2, value=it["descripcion"])
        ws.cell(row=fila, column=3, value=float(it["cantidad"]))
        ws.cell(row=fila, column=4, value=float(it["importe"]))
        fila += 1

    for col, ancho in [(1, 16), (2, 40), (3, 12), (4, 14)]:
        ws.column_dimensions[get_column_letter(col)].width = ancho

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_inventario_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    fila = _titulo(ws, "Valor de inventario", f"Generado {data['generado'][:10]}", 5)

    ws.cell(row=fila, column=1, value="Valor total (a costo)").font = Font(bold=True)
    ws.cell(row=fila, column=2, value=float(data["valor_total"]))
    ws.cell(row=fila + 1, column=1, value="Productos activos").font = Font(bold=True)
    ws.cell(row=fila + 1, column=2, value=data["productos_activos"])
    ws.cell(row=fila + 2, column=1, value="Sin stock").font = Font(bold=True)
    ws.cell(row=fila + 2, column=2, value=data["sin_stock"])
    ws.cell(row=fila + 3, column=1, value="Bajo mínimo").font = Font(bold=True)
    ws.cell(row=fila + 3, column=2, value=data["bajo_minimo"])

    fila += 5
    ws.cell(row=fila, column=1, value="PRODUCTOS QUE REQUIEREN REPOSICIÓN").font = Font(
        bold=True, color=ROJO
    )
    fila += 1
    for i, h in enumerate(["SKU", "Producto", "Stock", "Mínimo", "Estado"], start=1):
        ws.cell(row=fila, column=i, value=h)
    _estilar_cabecera(ws, fila, 5)
    fila += 1
    for a in data["alertas"]:
        ws.cell(row=fila, column=1, value=a["sku"])
        ws.cell(row=fila, column=2, value=a["nombre"])
        ws.cell(row=fila, column=3, value=float(a["stock_actual"]))
        ws.cell(row=fila, column=4, value=float(a["stock_minimo"]))
        ws.cell(
            row=fila,
            column=5,
            value="Sin stock" if a["estado"] == "SIN_STOCK" else "Bajo mínimo",
        )
        fila += 1

    for col, ancho in [(1, 16), (2, 40), (3, 12), (4, 12), (5, 14)]:
        ws.column_dimensions[get_column_letter(col)].width = ancho

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------
# PDF
# --------------------------------------------------------------------------
def _soles(v) -> str:
    return f"{Decimal(str(v)):,.2f}"


def exportar_ventas_pdf(data: dict) -> bytes:
    from weasyprint import HTML

    max_total = max((Decimal(d["total"]) for d in data["por_dia"]), default=Decimal("0")) or Decimal(
        "1"
    )
    barras = [
        {**d, "pct": float(Decimal(d["total"]) / max_total * 100)} for d in data["por_dia"]
    ]

    html = _env().get_template("reporte_ventas.html").render(
        empresa=EMPRESA,
        logo=_asset_data_url("logo_zonaxtrema.png"),
        data=data,
        barras=barras,
        soles=_soles,
        generado=hoy_local().isoformat(),
    )
    return HTML(string=html).write_pdf()
